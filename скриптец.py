import requests
from bs4 import BeautifulSoup
import openpyxl
from urllib.parse import urljoin, urlparse
from tqdm import tqdm

# Функция для проверки корректности URL
def validate_url(url):
    """
    Проверяет, имеет ли URL схему (http:// или https://).
    Если схема отсутствует, добавляет https://.
    """
    parsed = urlparse(url)
    if not parsed.scheme:
        url = 'https://' + url  # Добавляем https://, если схема отсутствует
    return url

# Функция для проверки кода ответа сервера с отключёнными перенаправлениями
def check_response(url):
    """
    Отправляет GET-запрос по URL с отключённым автоматическим перенаправлением (allow_redirects=False).
    Возвращает кортеж: (HTTP-код, содержимое страницы, значение заголовка "Location" если имеется)
    """
    try:
        url = validate_url(url)
        response = requests.get(url, timeout=10, allow_redirects=False)
        location = response.headers.get('Location', '')
        return response.status_code, response.content, location
    except requests.RequestException as e:
        print(f"Ошибка при запросе {url}: {e}")
        return None, None, None

# Функция для поиска всех внутренних ссылок на сайте
def find_internal_links(url, domain):
    """
    Находит внутренние ссылки на странице.
    Преобразует относительные ссылки в абсолютные и оставляет только те, у которых netloc совпадает с исходным доменом.
    """
    try:
        url = validate_url(url)
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        links = soup.find_all('a', href=True)
        internal_links = []
        for link in links:
            href = link['href']
            full_url = urljoin(url, href)
            # Сравниваем netloc исходного URL и полученной ссылки
            if urlparse(full_url).netloc == domain:
                internal_links.append(full_url)
        return internal_links
    except Exception as e:
        print(f"Ошибка при поиске внутренних ссылок на {url}: {e}")
        return []

# Функция для поиска ссылок на заданный домен
def find_target_links(url, target_domain):
    """
    Ищет на странице ссылки, в URL которых содержится указанный внешний домен.
    Возвращает список кортежей: (URL страницы, полный URL ссылки, текст ссылки).
    """
    try:
        url = validate_url(url)
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        links = soup.find_all('a', href=True)
        found_links = []
        for link in links:
            href = link['href']
            full_url = urljoin(url, href)
            if target_domain.lower() in urlparse(full_url).netloc.lower():
                anchor_text = link.get_text(strip=True)
                found_links.append((url, full_url, anchor_text))
        return found_links
    except Exception as e:
        print(f"Ошибка при поиске ссылок на {url}: {e}")
        return []

# Функция для записи данных в Excel
def write_to_excel(data, redirects, other_codes, output_file='result.xlsx'):
    workbook = openpyxl.Workbook()

    # Лист для найденных ссылок
    sheet_links = workbook.active
    sheet_links.title = 'Links'
    sheet_links.append(['URL страницы', 'Страница со ссылкой', 'Ссылка', 'Анкор'])
    for row in data:
        url_page, links_data = row
        for link_info in links_data:
            sheet_links.append([url_page, link_info[0], link_info[1], link_info[2]])

    # Лист для редиректов (301-307)
    sheet_redirects = workbook.create_sheet(title='Redirects (301-307)')
    sheet_redirects.append(['URL страницы', 'Код ответа', 'Куда перенаправляет'])
    for redirect_url, code, location in redirects:
        sheet_redirects.append([redirect_url, code, location])

    # Лист для других кодов ответа
    sheet_other_codes = workbook.create_sheet(title='Other Codes')
    sheet_other_codes.append(['URL страницы', 'Код ответа'])
    for url, code in other_codes:
        sheet_other_codes.append([url, code])

    workbook.save(output_file)
    print(f"Результаты сохранены в файл {output_file}")

# Функция для чтения URL из файла
def read_urls_from_file(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            urls = [line.strip() for line in file if line.strip()]
        return urls
    except FileNotFoundError:
        print(f"Файл {file_path} не найден.")
        return []

# Основная логика
def main(input_file, target_domain, output_file='result.xlsx'):
    input_urls = read_urls_from_file(input_file)
    if not input_urls:
        print("Нет URL для обработки.")
        return

    result_data = []
    redirects = []
    other_codes = []

    for url in tqdm(input_urls, desc="Обрабатываем URL", unit="URL"):
        status_code, content, location = check_response(url)
        if status_code is None:
            continue
        if status_code == 200:
            domain = urlparse(validate_url(url)).netloc
            internal_links = find_internal_links(url, domain)
            all_links_to_check = [url] + internal_links
            for link in all_links_to_check:
                links = find_target_links(link, target_domain)
                if links:
                    result_data.append((link, links))
        elif 301 <= status_code <= 307:
            print(f"Страница {url} перенаправлена ({status_code}).")
            redirects.append((url, status_code, location))
        else:
            print(f"Страница {url} вернула код {status_code}.")
            other_codes.append((url, status_code))

    write_to_excel(result_data, redirects, other_codes, output_file)

if __name__ == "__main__":
    input_file = 'input_urls.txt'
    target_domain = input("Введите искомый внешний домен (например, fayngor.ru): ").strip()
    main(input_file, target_domain)
